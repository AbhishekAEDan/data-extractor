"""
writers_it.py

Writes parsed IT lesson rows to formatted .xlsx workbooks. Layouts mirror the
Form 1 IT spreadsheet templates plus the "AI Sample Sheets/content
sheets.xlsx" examples (headers copied exactly, including their spacing/typo
quirks -- "Acitvity Intro ", "Module Title 1", "Think about it " etc.):

    Unit cover page & Learning Objective.xlsx
    Introduction Topic Overview.xlsx
    Key Concepts (Main Content).xlsx
    Practical Activity (Hands-On).xlsx
    Guided Practice (With Support).xlsx
    Independent Practice.xlsx
    Challenge Higher-Order Thinking.xlsx
    Knowledge Check (Quick Assessment).xlsx
    Common Mistakes & Tips.xlsx        (one row per Mistake/Tip pair)
    Real-World Application.xlsx
    Summary Key Takeaways.xlsx
    Answer Solution.xlsx               (answers fanned into numbered slots)
    Suggested Interactive Moments.xlsx (one row per suggested moment)
    full_extract.csv                   (debug: every parsed field, stays CSV)

Only the IT path is xlsx -- ELA (writers.py) still writes CSV through the
shared `write_csv` helper, which is left untouched.

All values are verbatim slices from the parser -- no rewriting happens here;
the only text the writer removes is its own internal `[TABLE-REQ: ...]` /
`[TABLE: ...]` wrappers.

Author: AbhishekAEDan
"""
__author__ = "AbhishekAEDan"

import csv
import os
import re

from parser_it import IT_ORDER
from writers import unit_sort_key, join_nonempty


def _sorted_rows(rows):
    out = [r for r in rows if r.get("_doc_type") == "it_lesson"]
    return sorted(out, key=lambda r: unit_sort_key(r.get("Module Number", "")))


# internal, non-textual row keys that must never become a sheet column
_PRIVATE_KEYS = {"_image_files"}


def order_cols(rows):
    cols = [c for c in IT_ORDER if any(c in r for r in rows)]
    for r in rows:
        for k in r:
            if k not in cols and k not in _PRIVATE_KEYS:
                cols.append(k)
    return cols


def write_full_extract(rows, out_dir):
    cols = order_cols(rows)
    path = os.path.join(out_dir, "full_extract.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in cols})
    return path


# image / diagram / video marker lines authored in the documents, plus the
# "[IMAGE...]" placeholders the parser inserts for embedded pictures
_IMG_LINE = re.compile(
    r'^\s*\[\s*(?:diagram|image|screenshot|photo|picture|video|sim\b|'
    r'animation|gif|use of a video)', re.I)


def split_images(text):
    """Split a section's verbatim text into (body, images). Image lines are
    the bracketed marker lines; a marker whose ']' falls on a later line is
    consumed up to that line."""
    body, imgs = [], []
    in_img = False
    for ln in (text or "").split("\n"):
        s = ln.strip()
        if in_img:
            imgs[-1] += "\n" + s
            if "]" in s:
                in_img = False
            continue
        if _IMG_LINE.match(s):
            imgs.append(s)
            in_img = "]" not in s
        else:
            body.append(ln)
    return "\n".join(body).strip(), "\n".join(imgs).strip()


# one-line table markers the parser emits for Word tables:
# "[TABLE: a | b || c | d]"
_TBL_LINE = re.compile(r'^\s*\[TABLE:\s?(.*)\]\s*$')


def split_tables(text):
    """Split a section's verbatim text into (body, tables). Table marker
    lines are pulled out and rendered for a spreadsheet cell: one row per
    line, cells joined with ' | '. Several tables in one section are
    separated by a blank line.

    NOTE: no sheet writes the table text any more -- the sheets carry only
    the yellow "TABLE REQUIRED" placeholder. This is kept because the parser
    still emits the [TABLE: ...] marker (it reaches full_extract.csv for
    debugging) and `_body` needs to drop it from every sheet cell."""
    body, tables = [], []
    for ln in (text or "").split("\n"):
        m = _TBL_LINE.match(ln)
        if m:
            tables.append("\n".join(r.strip() for r in m.group(1).split(" || ")))
        else:
            body.append(ln)
    return "\n".join(body).strip(), "\n\n".join(tables).strip()


def _is_img(line):
    """True for an image/diagram/screenshot/video marker line. Slot-splitting
    helpers use this so a marker is never counted as a paragraph, an answer,
    a mistake or a moment of its own -- it rides with the slot it sat in and
    is moved to that slot's adjacent image column at write time."""
    return bool(_IMG_LINE.match((line or "").strip()))


def _body(r, section):
    """One section's verbatim text with the [TABLE: ...] markers dropped.

    Image markers and the [TABLE-REQ: ...] placeholder deliberately STAY in
    the flow: the placeholder belongs in the paragraph cell, and each image
    marker has to survive slot assignment so it can be routed to the image
    column sitting immediately after its own slot."""
    return split_tables(r.get(section, ""))[0].strip()


class _C:
    """A content cell that owns the image markers found inside it.

    `text` is the cell's content with marker lines removed; `img` is the
    marker lines, newline-joined. `_finalize` turns each `_C` into its
    content column plus an image column placed immediately after it."""
    __slots__ = ("text", "img")

    def __init__(self, text):
        self.text, self.img = split_images(text or "")

    def __bool__(self):
        return bool(self.text or self.img)


def _finalize(path, header, rows):
    """Write a sheet whose row entries are plain strings or `_C` cells.

    Every `_C` position expands to its content column followed by an image
    column -- but only when at least one row actually carries markers there,
    the same dynamic philosophy the slot columns already use. A sheet that
    ends up with exactly one image column names it plain "Image"; with
    several, each is named "<content column> Image" so it is obvious which
    paragraph it belongs to. There is no trailing catch-all Image column."""
    keep = {i for i in range(len(header))
            if any(isinstance(r[i], _C) and r[i].img for r in rows)}
    single = len(keep) == 1
    out_h, out_rows = [], [[] for _ in rows]
    for i, name in enumerate(header):
        out_h.append(name)
        for j, r in enumerate(rows):
            v = r[i]
            out_rows[j].append(v.text if isinstance(v, _C) else v)
        if i in keep:
            out_h.append("Image" if single else f"{str(name).strip()} Image")
            for j, r in enumerate(rows):
                v = r[i]
                out_rows[j].append(v.img if isinstance(v, _C) else "")
    write_xlsx(path, out_h, out_rows)
    return path


# ---------- xlsx output ----------

# the parser's placeholder wrapper: "[TABLE-REQ: TABLE REQUIRED - see X.docx
# (Term 1)]". It rides in the paragraph flow (unlike "[TABLE: ...]", which is
# pulled out to the Table column) so the cell shows where the table belongs.
# The wrapper exists only so the matching helpers and this writer can spot the
# line unambiguously -- it is stripped on the way into the cell.
_TREQ_LINE = re.compile(r'^[ \t]*\[TABLE-REQ:[ ]?(.*?)\][ \t]*$', re.M)

_TREQ_TEXT = "TABLE REQUIRED"

_YELLOW = "FFFF00"

# narrow columns: identifiers and titles. Everything else is content.
_NARROW_RE = re.compile(r'(number|no\.|title|section |term|form)', re.I)


def _clean_cell(value):
    """Cell text on its way into the sheet: the internal [TABLE-REQ: ...]
    wrapper is unwrapped to its plain text, and any [TABLE: ...] marker line
    is removed outright -- table text is never written to a sheet, only the
    yellow placeholder is. The second pass is defensive; `_body` already
    drops table markers before slot splitting."""
    if not isinstance(value, str):
        return value
    if "[TABLE-REQ:" in value:
        value = _TREQ_LINE.sub(lambda m: m.group(1), value)
    if "[TABLE:" in value:
        value = "\n".join(ln for ln in value.split("\n")
                          if not _TBL_LINE.match(ln)).strip()
    return value


def _is_treq(line):
    """True for the parser's table placeholder line. Matching helpers use it
    to make sure the placeholder is never mistaken for a heading, a mistake,
    a tip or an answer -- it is plain content that rides where it sat."""
    return bool(_TREQ_LINE.match((line or "").strip()))


def _col_width(name):
    if str(name).strip().lower().endswith("image"):
        return 30
    return 18 if _NARROW_RE.search(str(name)) else 60


# doc-wide index of extracted image files, built by write_all_it:
#   marker text -> [absolute file path, ...]  (consumed in order)
_IMAGE_INDEX = {}


def _index_images(rows, out_dir):
    """Write every embedded picture the parser captured to
    <out_dir>/_images/<docfile>_<n><ext> and index the files by the marker
    line that represents them, so any sheet's Image cell can be matched back
    to its picture without changing the per-sheet writer signatures."""
    _IMAGE_INDEX.clear()
    img_dir = os.path.join(out_dir, "_images")
    for r in rows:
        recs = r.get("_image_files") or []
        if not recs:
            continue
        os.makedirs(img_dir, exist_ok=True)
        stem = os.path.splitext(r.get("_file", "doc"))[0]
        stem = re.sub(r'[^\w\- ]+', "_", stem)
        for n, rec in enumerate(recs, 1):
            path = os.path.join(img_dir, f"{stem}_{n}{rec.get('ext') or '.png'}")
            try:
                with open(path, "wb") as fh:
                    fh.write(rec["blob"])
            except OSError:
                continue
            _IMAGE_INDEX.setdefault(rec["marker"], []).append(path)


def _cell_image_files(text):
    """Image files whose marker appears in this Image cell, consumed so the
    same picture is not embedded twice when markers repeat across rows."""
    out, t = [], text or ""
    for marker, paths in _IMAGE_INDEX.items():
        for _ in range(min(t.count(marker), len(paths))):
            out.append(paths.pop(0))
    return out


_MAX_IMG_PX = 200


def _embed_images(ws, row_idx, col_idx, paths):
    """Anchor pictures inside one cell, scaled to <=200px wide, stacked, and
    grow the row so they roughly fit. Returns the stack height in pixels."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    offset = 0
    for p in paths:
        try:
            pic = XLImage(p)
        except Exception:
            continue
        w, h = pic.width, pic.height
        if w > _MAX_IMG_PX:
            h = max(1, int(h * _MAX_IMG_PX / float(w)))
            w = _MAX_IMG_PX
        pic.width, pic.height = w, h
        pic.anchor = OneCellAnchor(
            _from=AnchorMarker(col=col_idx - 1, row=row_idx - 1,
                               colOff=0, rowOff=pixels_to_EMU(offset)),
            ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h)))
        ws.add_image(pic)
        offset += h + 4
    return offset


def write_xlsx(path, header, data_rows, sheet_title=None):
    """Write one IT template sheet as a formatted .xlsx.

    Formatting: bold header row, frozen top row, every cell wrapped and
    top-aligned, identifier/title columns narrow and content columns wide.
    The first column (the module/lesson number) is written as TEXT-formatted
    string so Excel keeps "3.10" instead of coercing it to 3.1 -- this is the
    xlsx replacement for the `="3.10"` CSV hack in writers.write_csv.
    Any cell containing "TABLE REQUIRED" gets a solid yellow fill."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    if sheet_title:
        ws.title = str(sheet_title)[:31]
    wrap = Alignment(wrap_text=True, vertical="top")
    head_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    bold = Font(bold=True)
    yellow = PatternFill(fill_type="solid", fgColor=_YELLOW, start_color=_YELLOW,
                         end_color=_YELLOW)

    ws.append(list(header))
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.alignment = head_align
        ws.column_dimensions[get_column_letter(c)].width = _col_width(header[c - 1])

    # image columns now sit next to the paragraph they belong to, so there can
    # be several of them per sheet: "Image" or "<content column> Image"
    img_cols = [c for c in range(1, len(header) + 1)
                if str(header[c - 1]).strip() == "Image"
                or str(header[c - 1]).strip().endswith(" Image")]

    for i, r in enumerate(data_rows, start=2):
        for c in range(1, len(header) + 1):
            v = r[c - 1] if c - 1 < len(r) else ""
            v = _clean_cell(v if v is not None else "")
            cell = ws.cell(row=i, column=c, value=v)
            cell.alignment = wrap
            if c == 1:
                # keep "3.10" as text -- never let Excel see it as a number
                cell.number_format = "@"
            if isinstance(v, str) and _TREQ_TEXT in v:
                cell.fill = yellow
        for img_col in img_cols:
            files = _cell_image_files(r[img_col - 1] if img_col - 1 < len(r) else "")
            if files:
                px = _embed_images(ws, i, img_col, files)
                if px:
                    ws.row_dimensions[i].height = max(
                        ws.row_dimensions[i].height or 0, px * 0.75 + 6)

    ws.freeze_panes = "A2"
    wb.save(path)


def _bare(s):
    """Copy of a line with list markers/indent removed -- for MATCHING only;
    cell content keeps the markers."""
    return s.lstrip("\t •·●▪-–—")


def _mod(r):
    return r.get("Module Number", "")


def _title(r):
    return r.get("Module Title", "")


# ---------- simple one-block sheets ----------

def write_cover(rows, out_dir):
    header = ["Module Number", "Module Title 1", "Module Descriptive Title",
              "Learning Objectives"]
    data = []
    for r in _sorted_rows(rows):
        obj = _C(_body(r, "Learning Objectives"))
        if obj:
            data.append([_mod(r), _title(r),
                         r.get("Module Descriptive Title", ""), obj])
    return _finalize(os.path.join(
        out_dir, "Unit cover page & Learning Objective.xlsx"), header, data)


def pad(lst, n):
    """Return lst padded with empty strings to length n."""
    return list(lst) + [""] * (n - len(lst))


def pad_cells(lst, n):
    """pad() for slots holding `_C` cells -- pads with empty ones so the
    content/image column pair stays aligned on short rows."""
    return list(lst) + [_C("") for _ in range(n - len(lst))]


def write_intro(rows, out_dir):
    """Introduction paragraphs fan into one slot each. The number of
    Paragraph columns is the maximum paragraph count across all documents;
    shorter rows pad with ''. An image marker is not a paragraph -- it joins
    the paragraph above it and surfaces in that slot's image column."""
    parsed = []
    for r in _sorted_rows(rows):
        text = _body(r, "Introduction / Topic Overview")
        if not text:
            continue
        paras = []                      # each entry = [lines of one paragraph]
        for ln in text.split("\n"):
            if not ln.strip():
                continue
            if _is_img(ln) and paras:
                paras[-1].append(ln)
            else:
                paras.append([ln])
        parsed.append((r, [_C("\n".join(p)) for p in paras]))
    n = max([len(p) for _, p in parsed] + [1])
    header = (["Module Number", "Module Title 1"]
              + [f"Paragraph {i}" for i in range(1, n + 1)])
    data = [[_mod(r), _title(r)] + pad_cells(paras, n)
            for r, paras in parsed]
    return _finalize(os.path.join(
        out_dir, "Introduction Topic Overview.xlsx"), header, data)


_SUBHEAD_RE = re.compile(r'^\d{1,2}\.\d{1,2}\b')


def write_key_concepts(rows, out_dir):
    """Intro text before the first N.N sub-heading goes to 'Key concepts
    intro'; each N.N block splits into its Heading slot (the '3.3 Title'
    line) and its Paragraph slot (the text that follows the heading). The
    number of Heading/Paragraph pairs is the maximum block count across all
    documents; shorter rows pad with ''.

    An image marker inside an N.N block belongs to that block's paragraph, so
    it lands in the image column immediately after that Paragraph N."""
    parsed = []
    for r in _sorted_rows(rows):
        text = _body(r, "Key Concepts")
        if not text:
            continue
        intro, blocks = [], []          # block = [heading, [body lines]]
        for ln in text.split("\n"):
            # a marker line is never a sub-heading; _SUBHEAD_RE cannot match
            # one, so it simply accumulates into the block it sat in
            if _SUBHEAD_RE.match(_bare(ln).strip()):
                blocks.append([ln.strip(), []])
            elif blocks:
                blocks[-1][1].append(ln)
            else:
                intro.append(ln)
        parsed.append((r, _C("\n".join(intro)),
                       [(h, _C("\n".join(b))) for h, b in blocks]))
    n = max([len(b) for _, _, b in parsed] + [1])
    header = ["Lesson No.", "Key concepts intro"]
    for i in range(1, n + 1):
        header += [f"Heading {i}", f"Paragraph {i}"]
    data = []
    for r, intro, blocks in parsed:
        cells = []
        for h, b in blocks:
            cells += [h, b]
        # pad in Heading/Paragraph pairs so the Paragraph slots keep holding
        # `_C` cells and their image columns stay aligned
        while len(cells) < n * 2:
            cells += ["", _C("")]
        data.append([_mod(r), intro] + cells)
    return _finalize(os.path.join(
        out_dir, "Key Concepts (Main Content).xlsx"), header, data)


# ---------- practical activity ----------

# bold sub-labels inside the Practical Activity section -> output column
_PA_BUCKETS = [
    (re.compile(r'^materials\b', re.I),                 "steps"),
    (re.compile(r'^instructions?\b', re.I),             "steps"),
    (re.compile(r'^(guide|steps?)\b\s*[:/]?', re.I),    "steps"),
    (re.compile(r'^step\s*\d', re.I),                   "steps"),
    (re.compile(r'^success criteria\b', re.I),          "success"),
    (re.compile(r'^step\s*\d+\s*[:.]?\s*review\b', re.I), "success"),
    (re.compile(r'^discussion\b', re.I),                "discussion"),
    (re.compile(r'^troubleshooting\b', re.I),           "trouble"),
    (re.compile(r'^.{0,20}rubric\b', re.I),             "rubric"),
]

_PA_REVIEW = re.compile(r'^step\s*\d+\s*[:.]?\s*review\b', re.I)


def _pa_bucket(line):
    # a marker line is content, never a bucket label -- guarded explicitly
    # because the loose "^.{0,20}rubric" pattern could otherwise be tripped
    # by something like "[SCREENSHOT: A rubric ...]"
    if _is_img(line) or _is_treq(line):
        return None
    if _PA_REVIEW.match(line):
        return "success"
    for rx, b in _PA_BUCKETS:
        if rx.match(line):
            return b
    return None


def write_practical(rows, out_dir):
    """Each bold sub-label starts a bucket; every bucket gets its own image
    column immediately after it when any document has markers there."""
    header = ["Module Number", "Module Title 1", "Practical Title",
              "Acitvity Intro ", "Guide/ Steps", "Success Criteria",
              "Discussion", "Troubleshooting", "Rubric "]
    data = []
    for r in _sorted_rows(rows):
        text = _body(r, "Practical Activity")
        if not text:
            continue
        buckets = {"intro": [], "steps": [], "success": [],
                   "discussion": [], "trouble": [], "rubric": []}
        title = r.get("Practical Title", "")
        cur = "intro"
        for ln in text.split("\n"):
            s = ln.strip()
            if not s:
                continue
            # "Task: ..." lead names the activity when the heading didn't
            if not title and re.match(r'^task\s*[:—–-]', _bare(s), re.I):
                title = re.split(r'[:—–-]', s, 1)[1].strip().split(". ")[0]
            # a marker never opens a bucket -- _pa_bucket cannot match one, so
            # it stays with the bucket it appeared in
            b = _pa_bucket(_bare(s))
            if b:
                cur = b
            buckets[cur].append(s)
        data.append([_mod(r), _title(r), title]
                    + [_C("\n".join(buckets[k])) for k in
                       ("intro", "steps", "success", "discussion",
                        "trouble", "rubric")])
    return _finalize(os.path.join(
        out_dir, "Practical Activity (Hands-On).xlsx"), header, data)


# ---------- practice sheets (top paragraph + body) ----------

def _split_top(text, body_re):
    """Leading lines before the first body line (scenario/question/numbered)
    form the top paragraph; the rest is the body, verbatim. A marker line can
    never start the body (body_re cannot match one), so it rides with whichever
    part it sat in and surfaces in that part's own image column."""
    top, body = [], []
    for ln in text.split("\n"):
        s = ln.strip()
        if not s:
            continue
        if not body and not body_re.match(_bare(s)):
            top.append(s)
        else:
            body.append(s)
    return "\n".join(top).strip(), "\n".join(body).strip()


_GP_BODY = re.compile(r'^(\d{1,2}[.)]\s|scenario\b|situation\b)', re.I)
_IP_BODY = re.compile(r'^(\d{1,2}[.)]\s|[a-e][.)]\s|[A-E]\.\s)', re.I)
_KC_BODY = re.compile(r'^\d{1,2}[.)]?\s|^what\b|^which\b|^why\b|^who\b|^how\b',
                      re.I)


def write_guided(rows, out_dir):
    header = ["Module Number", "Module Title 1", "Top paragraph",
              "Practice Scenarios"]
    data = []
    for r in _sorted_rows(rows):
        text = _body(r, "Guided Practice")
        if not text:
            continue
        top, body = _split_top(text, _GP_BODY)
        data.append([_mod(r), _title(r), _C(top), _C(body)])
    return _finalize(os.path.join(
        out_dir, "Guided Practice (With Support).xlsx"), header, data)


def write_independent(rows, out_dir):
    header = ["Module Number", "Module Title 1", "Top paragraph",
              "Independent Practice "]
    data = []
    for r in _sorted_rows(rows):
        text = _body(r, "Independent Practice")
        if not text:
            continue
        top, body = _split_top(text, _IP_BODY)
        data.append([_mod(r), _title(r), _C(top), _C(body)])
    return _finalize(os.path.join(
        out_dir, "Independent Practice.xlsx"), header, data)


def write_challenge(rows, out_dir):
    header = ["Module Number", "Module Title 1", "Section Content"]
    data = []
    for r in _sorted_rows(rows):
        obj = _C(_body(r, "Challenge / Higher-Order Thinking"))
        if obj:
            data.append([_mod(r), _title(r), obj])
    return _finalize(os.path.join(
        out_dir, "Challenge Higher-Order Thinking.xlsx"), header, data)


def write_knowledge_check(rows, out_dir):
    header = ["Module Number", "Module Title 1", "Instruction", "Questions"]
    data = []
    for r in _sorted_rows(rows):
        text = _body(r, "Knowledge Check")
        if not text:
            continue
        top, body = _split_top(text, _KC_BODY)
        data.append([_mod(r), _title(r), _C(top), _C(body)])
    return _finalize(os.path.join(
        out_dir, "Knowledge Check (Quick Assessment).xlsx"), header, data)


# ---------- common mistakes ----------

_MISTAKE_RE = re.compile(r'^mistake\s*[:—–-]?\s*', re.I)
_TIP_RE = re.compile(r'^tip\s*[:—–-]?\s*', re.I)


def write_mistakes(rows, out_dir):
    """One row per Mistake/Tip pair. Text that carries no explicit tags is
    kept whole in the Mistake column so nothing is dropped. An image marker is
    never a Mistake or a Tip of its own -- it rides with the part it sat in and
    lands in that column's own adjacent image column."""
    header = ["Module Number", "Module Title 1", "Mistake ", "Tip"]
    data = []
    for r in _sorted_rows(rows):
        text = _body(r, "Common Mistakes & Tips")
        if not text.strip():
            continue
        pairs = [["", ""]]
        cur = None
        tagged = False
        for ln in text.split("\n"):
            s = ln.strip()
            if not s:
                continue
            if _MISTAKE_RE.match(_bare(s)):
                tagged = True
                if pairs[-1][0]:
                    pairs.append(["", ""])
                pairs[-1][0] = s
                cur = 0
            elif _TIP_RE.match(_bare(s)):
                tagged = True
                pairs[-1][1] = s
                cur = 1
            elif cur is not None:
                pairs[-1][cur] = (pairs[-1][cur] + "\n" + s).strip()
            else:
                pairs[-1][0] = join_nonempty(pairs[-1][0], s)
        if not tagged and pairs == [["", ""]]:
            pairs = [[text.strip(), ""]]
        for m, t in pairs:
            if m or t:
                data.append([_mod(r), _title(r), _C(m), _C(t)])
    return _finalize(os.path.join(
        out_dir, "Common Mistakes & Tips.xlsx"), header, data)


def write_real_world(rows, out_dir):
    header = ["Module Number", "Module Title 1", "Intro/ paragraph", "Points",
              "Think about it "]
    data = []
    for r in _sorted_rows(rows):
        text = _body(r, "Real-World Application")
        if not text:
            continue
        # a trailing "Think about it"/question block splits off when marked
        think = ""
        m = re.search(r'^[\s•·●▪\t-]*think about it\b.*$', text, re.I | re.M)
        if m:
            text, think = text[:m.start()].rstrip(), text[m.start():].strip()
        data.append([_mod(r), _title(r), _C(text), "", _C(think)])
    return _finalize(os.path.join(
        out_dir, "Real-World Application.xlsx"), header, data)


def write_summary(rows, out_dir):
    header = ["Module Number", "Module Title 1", "Summary Content "]
    data = []
    for r in _sorted_rows(rows):
        obj = _C(_body(r, "Summary / Key Takeaways"))
        if obj:
            data.append([_mod(r), _title(r), obj])
    return _finalize(os.path.join(
        out_dir, "Summary Key Takeaways.xlsx"), header, data)


# ---------- answers & solutions ----------

# sub-blocks inside the Answers & Solutions section, matched by name
_ANS_BLOCKS = [
    (re.compile(r'^guided practice\b.*[:]?', re.I),      "gp"),
    (re.compile(r'^independent practice\b.*[:]?', re.I), "ip"),
    (re.compile(r'^(challenge|higher-order|evaluate a scenario)\b.*', re.I), "hot"),
    (re.compile(r'^knowledge check\b.*[:]?', re.I),      "kc"),
]

_NUMLINE = re.compile(r'^(\d{1,2})[.)]\s*(.*)$')


def _fan(lines):
    """Fan answer lines into items. Explicit '1.' numbering wins; unnumbered
    lines each start a new item. No slot cap -- callers size and pad their
    own columns."""
    items = []
    for ln in lines:
        # the table placeholder and image markers are content, never answers
        # of their own -- they ride with the answer they followed so the
        # numbered slots don't shift. The marker is separated out into the
        # slot's adjacent image column later, by `_C`.
        if (_is_treq(ln) or _is_img(ln)) and items:
            items[-1].append(ln)
            continue
        m = _NUMLINE.match(_bare(ln).strip())
        if m:
            items.append([m.group(2).strip()])
        elif items and re.match(r'^[a-e][).]\s', _bare(ln).strip(), re.I) \
                and not re.match(r'^[a-e][).]\s', _bare(items[-1][0]), re.I):
            # a)-e) option lines ride with the question above them -- unless
            # the previous item is itself a lettered answer (knowledge-check
            # style: each "b) ..." line is its own answer)
            items[-1].append(ln)
        elif items and _bare(ln).strip().lower().startswith(("reason", "fix", "explanation")):
            items[-1].append(ln)
        elif ln.strip():
            items.append([ln.strip()])
    return ["\n".join(x).strip() for x in items]


def write_answers(rows, out_dir):
    """Answers fanned into numbered slots per category. Each answer slot gets
    its own adjacent image column when any document has a marker in it."""
    parsed = []
    for r in _sorted_rows(rows):
        text = _body(r, "Answers & Solutions")
        if not text.strip():
            continue
        buckets = {"gp": [], "ip": [], "hot": [], "kc": [], "": []}
        cur = ""
        for ln in text.split("\n"):
            s = ln.strip()
            if not s:
                continue
            hit = None
            for rx, b in _ANS_BLOCKS:
                if rx.match(_bare(s)) and len(s) <= 60:
                    hit = b
                    break
            if hit:
                # "Evaluate a Scenario:" is a sub-part of the challenge block
                if hit == "hot" and cur == "hot":
                    buckets["hot"].append(s)
                else:
                    cur = hit
                continue
            buckets[cur].append(s)
        # answers before any recognised block label ride on Guided Practice
        if buckets[""]:
            buckets["gp"] = buckets[""] + buckets["gp"]
        parsed.append((r, {k: [_C(x) for x in _fan(buckets[k])]
                           for k in ("gp", "ip", "hot", "kc")}))
    cats = [("gp", "Guided Practice"), ("ip", "Independent Practice"),
            ("hot", "Higher-Order Thinking"), ("kc", "Knowledge Check")]
    widths = {k: max([len(f[k]) for _, f in parsed] + [1]) for k, _ in cats}
    header = ["Module Number", "Module Title 1"]
    for k, label in cats:
        header += [f"{label} {i}" for i in range(1, widths[k] + 1)]
    data = []
    for r, fans in parsed:
        cells = []
        for k, _ in cats:
            cells += pad_cells(fans[k], widths[k])
        data.append([_mod(r), _title(r)] + cells)
    return _finalize(os.path.join(
        out_dir, "Answer Solution.xlsx"), header, data)


# ---------- suggested interactive moments ----------

_MOMENT_START = re.compile(r'^(moment\s*\d+|location\s*[:—–-]|where it fits\s*[:—–-])',
                           re.I)
_SECTION_LINE = re.compile(r'^(location|where it fits)\s*[:—–-]\s*(.*)$', re.I)


def write_moments(rows, out_dir):
    """One row per suggested moment. Groups start at a 'Moment N' or
    'Location:'/'Where it fits:' line; the Section column carries the
    location, the Suggestion column the whole group verbatim."""
    header = ["Module title ", "Section ", "Suggestion"]
    data = []
    for r in _sorted_rows(rows):
        text = _body(r, "Suggested Interactive Moments")
        if not text.strip():
            continue
        groups = []
        for ln in text.split("\n"):
            s = ln.strip()
            if not s:
                continue
            # a new group starts at "Moment N", or at a Location line when
            # the current group already has one (bullet-list style docs)
            # a marker line never opens a group -- it belongs to the moment
            # it was written under
            if _is_img(s) and groups:
                groups[-1].append(s)
                continue
            is_loc = bool(_SECTION_LINE.match(_bare(s)))
            starts_new = bool(re.match(r'^moment\s*\d+', _bare(s), re.I)) or (
                is_loc and groups and any(_SECTION_LINE.match(_bare(x))
                                          for x in groups[-1]))
            if starts_new or not groups:
                groups.append([s])
            else:
                groups[-1].append(s)
        mod_title = join_nonempty(_mod(r), _title(r), sep=" ")
        for g in groups:
            section = ""
            for x in g:
                m = _SECTION_LINE.match(_bare(x))
                if m:
                    section = m.group(2).strip()
                    break
            data.append([mod_title, section, _C("\n".join(g))])
    return _finalize(os.path.join(
        out_dir, "Suggested Interactive Moments.xlsx"), header, data)


def write_all_it(rows, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    # embedded pictures land on disk once, then get anchored into whichever
    # sheet's Image cell carries their marker
    _index_images(rows, out_dir)
    paths = [
        write_full_extract(rows, out_dir),
        write_cover(rows, out_dir),
        write_intro(rows, out_dir),
        write_key_concepts(rows, out_dir),
        write_practical(rows, out_dir),
        write_guided(rows, out_dir),
        write_independent(rows, out_dir),
        write_challenge(rows, out_dir),
        write_knowledge_check(rows, out_dir),
        write_mistakes(rows, out_dir),
        write_real_world(rows, out_dir),
        write_summary(rows, out_dir),
        write_answers(rows, out_dir),
        write_moments(rows, out_dir),
    ]
    return paths
