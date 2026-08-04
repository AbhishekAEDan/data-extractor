"""
xlsx_out.py

Shared .xlsx writing machinery for both subject paths (`writers.py` for ELA and
`writers_it.py` for IT). Everything here is presentation only -- it never
rewrites extracted text; the only strings it removes are the parser's own
internal `[TABLE-REQ: ...]` / `[TABLE: ...]` wrappers.

Formatting contract (identical for both subjects):
  * bold, left-aligned, wrapped header row; frozen at "A2"
  * every data cell wrapped and top-aligned
  * identifier/title columns 18 wide, image columns 30, content columns 60
  * column 1 (the unit/module number) written as a TEXT-formatted cell so
    Excel keeps "3.10" instead of coercing it to 3.1 -- this replaces the old
    `="3.10"` CSV hack that `writers.write_csv` used
  * any cell containing "TABLE REQUIRED" gets a solid yellow fill

Embedded pictures are optional: `_index_images` builds the marker -> file index
consumed by `write_xlsx`. A path that never calls it (ELA has no embedded-image
pipeline) simply gets no images -- the index stays empty and the embedding step
is a no-op.

Author: AbhishekAEDan
"""
__author__ = "AbhishekAEDan"

import os
import re

# one-line table markers the parser emits for Word tables:
# "[TABLE: a | b || c | d]" -- debug only, stripped from every sheet cell
_TBL_LINE = re.compile(r'^\s*\[TABLE:\s?(.*)\]\s*$')

# the parser's placeholder wrapper: "[TABLE-REQ: TABLE REQUIRED - see X.docx
# (Term 1)]". It rides in the paragraph flow (unlike "[TABLE: ...]") so the
# cell shows where the table belongs. The wrapper exists only so the matching
# helpers and this writer can spot the line unambiguously -- it is stripped on
# the way into the cell.
_TREQ_LINE = re.compile(r'^[ \t]*\[TABLE-REQ:[ ]?(.*?)\][ \t]*$', re.M)

_TREQ_TEXT = "TABLE REQUIRED"

_YELLOW = "FFFF00"

# light red -- "needs a human look". Used by the IT Answer Solution sheet to
# mark rows whose answers could not be paired with questions with confidence.
_RED = "FFC7CE"

# narrow columns: identifiers and titles. Everything else is content.
_NARROW_RE = re.compile(r'(number|no\.|title|section |term|form|document)', re.I)


def _clean_cell(value):
    """Cell text on its way into the sheet: the internal [TABLE-REQ: ...]
    wrapper is unwrapped to its plain text, and any [TABLE: ...] marker line
    is removed outright -- table text is never written to a sheet, only the
    yellow placeholder is. The second pass is defensive; the IT writer's
    `_body` already drops table markers before slot splitting."""
    if not isinstance(value, str):
        return value
    if "[TABLE-REQ:" in value:
        value = _TREQ_LINE.sub(lambda m: m.group(1), value)
    if "[TABLE:" in value:
        value = "\n".join(ln for ln in value.split("\n")
                          if not _TBL_LINE.match(ln)).strip()
    return value


def _col_width(name):
    if str(name).strip().lower().endswith("image"):
        return 30
    return 18 if _NARROW_RE.search(str(name)) else 60


# doc-wide index of extracted image files, built by _index_images:
#   marker text -> [absolute file path, ...]  (consumed in order)
_IMAGE_INDEX = {}


def _index_images(rows, out_dir):
    """Write every embedded picture the parser captured to
    <out_dir>/_images/<docfile>_<n><ext> and index the files by the marker
    line that represents them, so any sheet's Image cell can be matched back
    to its picture without changing the per-sheet writer signatures.

    Optional: a subject path with no embedded-image pipeline never calls this,
    which leaves the index empty and makes embedding a no-op."""
    _IMAGE_INDEX.clear()
    img_dir = os.path.join(out_dir, "_images")
    for r in rows:
        recs = r.get("_image_files") or []
        if not recs:
            continue
        os.makedirs(img_dir, exist_ok=True)
        stem = os.path.splitext(r.get("_file", "doc"))[0]
        stem = re.sub(r'[^\w\- ]+', "_", stem)
        for n, rec in enumerate(recs, 1):
            path = os.path.join(img_dir, f"{stem}_{n}{rec.get('ext') or '.png'}")
            try:
                with open(path, "wb") as fh:
                    fh.write(rec["blob"])
            except OSError:
                continue
            _IMAGE_INDEX.setdefault(rec["marker"], []).append(path)


def _cell_image_files(text):
    """Image files whose marker appears in this Image cell, consumed so the
    same picture is not embedded twice when markers repeat across rows."""
    if not _IMAGE_INDEX:
        return []
    out, t = [], text or ""
    for marker, paths in _IMAGE_INDEX.items():
        for _ in range(min(t.count(marker), len(paths))):
            out.append(paths.pop(0))
    return out


_MAX_IMG_PX = 200


def _embed_images(ws, row_idx, col_idx, paths):
    """Anchor pictures inside one cell, scaled to <=200px wide, stacked, and
    grow the row so they roughly fit. Returns the stack height in pixels."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    offset = 0
    for p in paths:
        try:
            pic = XLImage(p)
        except Exception:
            continue
        w, h = pic.width, pic.height
        if w > _MAX_IMG_PX:
            h = max(1, int(h * _MAX_IMG_PX / float(w)))
            w = _MAX_IMG_PX
        pic.width, pic.height = w, h
        pic.anchor = OneCellAnchor(
            _from=AnchorMarker(col=col_idx - 1, row=row_idx - 1,
                               colOff=0, rowOff=pixels_to_EMU(offset)),
            ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h)))
        ws.add_image(pic)
        offset += h + 4
    return offset


def write_xlsx(path, header, data_rows, sheet_title=None, fills=None):
    """Write one template sheet as a formatted .xlsx.

    Formatting: bold header row, frozen top row, every cell wrapped and
    top-aligned, identifier/title columns narrow and content columns wide.
    The first column (the unit/module number) is written as a TEXT-formatted
    string so Excel keeps "3.10" instead of coercing it to 3.1 -- this is the
    xlsx replacement for the `="3.10"` CSV hack the ELA writer used to need.
    Any cell containing "TABLE REQUIRED" gets a solid yellow fill.

    `fills` is an optional {(data_row_index, column_index): "RRGGBB"} map of
    extra solid fills, both indexes ZERO-BASED against `data_rows` / `header`.
    It is the one hook a caller has for marking individual cells (the IT
    Answer Solution sheet uses it for its light-red manual-check flag); every
    other sheet passes nothing and is unaffected.

    Embedded pictures are only looked up when a marker index has been built
    (see `_index_images`); with an empty index the image pass does nothing."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    if sheet_title:
        ws.title = str(sheet_title)[:31]
    wrap = Alignment(wrap_text=True, vertical="top")
    head_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    bold = Font(bold=True)
    yellow = PatternFill(fill_type="solid", fgColor=_YELLOW, start_color=_YELLOW,
                         end_color=_YELLOW)

    ws.append(list(header))
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.alignment = head_align
        ws.column_dimensions[get_column_letter(c)].width = _col_width(header[c - 1])

    # image columns sit next to the paragraph they belong to, so there can be
    # several of them per sheet: "Image" or "<content column> Image". Sheets
    # with no image columns at all (every ELA sheet) skip the pass entirely.
    # the unit/module number column must stay TEXT-formatted. It is column 1
    # on the ELA sheets and column 2 on the IT sheets, which lead with the
    # plain "Document" column.
    num_col = 2 if str(header[0]).strip() == "Document" else 1

    img_cols = [c for c in range(1, len(header) + 1)
                if str(header[c - 1]).strip() == "Image"
                or str(header[c - 1]).strip().endswith(" Image")]

    for i, r in enumerate(data_rows, start=2):
        for c in range(1, len(header) + 1):
            v = r[c - 1] if c - 1 < len(r) else ""
            v = _clean_cell(v if v is not None else "")
            cell = ws.cell(row=i, column=c, value=v)
            cell.alignment = wrap
            if c == num_col:
                # keep "3.10" as text -- never let Excel see it as a number
                cell.number_format = "@"
            if isinstance(v, str) and _TREQ_TEXT in v:
                cell.fill = yellow
            colour = (fills or {}).get((i - 2, c - 1))
            if colour:
                cell.fill = PatternFill(fill_type="solid", fgColor=colour,
                                        start_color=colour, end_color=colour)
        for img_col in img_cols:
            files = _cell_image_files(r[img_col - 1] if img_col - 1 < len(r) else "")
            if files:
                px = _embed_images(ws, i, img_col, files)
                if px:
                    ws.row_dimensions[i].height = max(
                        ws.row_dimensions[i].height or 0, px * 0.75 + 6)

    ws.freeze_panes = "A2"
    wb.save(path)
